"""
    Translates the spreadsheet - easy and quick.          (tested in Python 3.10)
     by Aleksandrov M. 2022 in Amsterdam

    How to use? That's how:

    >> ./spreadsheet_translate.py -n file.xlsx -c "nl" -t "en"               Does everything but via the terminal

    Args:  -n   name of the file
           -c   current language of the spreadsheet (e.g. nl, en)
           -t   target language of the spreadsheet

    from spreadsheet_translate import Spreadsheet                            import the directory

    spreadsheet = Spreadsheet("file.xlsx", current="nl", target="en")        Open the workbook
    spreadsheet.translate()                                                  Translate all the spreadsheets
    spreadsheet.save()                                                       Save the workbook in the new file
    
    IMPORTANT: The translation is possible due to the free directory "translate", 
    if your workbook/spreadsheet is too big you will run into an issue of 
    "MYMEMORY WARNING: YOU USED ALL AVAILABLE FREE TRANSLATIONS FOR TODAY". So, to avoid it
    try to use a paid service and slightly change the code. Personally, I use API from Microsoft Azure,
    very easy to integrate into the code (±5 minutes). I have made commentaries inside the code,
    to show where the translation itself happens.
"""

import logging
import argparse
from openpyxl import load_workbook
from translate import Translator


LEVEL = logging.INFO
FMT = '[%(levelname)s] %(asctime)s - %(message)s'
logging.basicConfig(level=LEVEL, format=FMT)

parser = argparse.ArgumentParser()
parser.add_argument("-n", "--name", type=str, required=True, help="The name of the Excel file")
parser.add_argument("-c", "--current", type=str, required=True, help="Current language of the spreadsheet")
parser.add_argument("-t", "--to", type=str, required=True, help="Target language of the spreadsheet")
args = parser.parse_args()


class Spreadsheet:

    """Opens, translates and saves the workbook (.xlsx)"""

    def __init__(self, filename_: str, current: str, target: str):
        self.translator = Translator(from_lang=current, to_lang=target)           #THIS IS THE PART THAT TRANSLATES THE TEXT    (1/2)
        self.workbook = None
        self.filename = filename_

    def translate_cell(self, text: str):

        """Translates individual cells after selecting non-numerical cells"""

        try:
            float(text)
        except ValueError:
            if text[0] != "=":
                return self.translator.translate(text)            #THIS IS THE PART THAT TRANSLATES THE TEXT      (2/2)
            return text
        except TypeError:
            return ''
        else:
            return round(float(text), 2)

    def transl_sheet(self, sheet: any):

        """Iterates through each cell in each row in each column and extracts its value"""

        logging.info('Worksheet "%s" - Translation started. Processing...', sheet.title)
        for col_n, column in enumerate(sheet.iter_cols(values_only=True), start=1):
            for row_n, cell in enumerate(column, start=1):
                try:
                    sheet.cell(row=row_n, column=col_n).value = self.translate_cell(cell)
                except AttributeError:
                    pass
        logging.info('Worksheet "%s" - Translation completed. Moving on...', sheet.title)
        return sheet

    def translate(self):

        """Loads the workbook and iterates through each spreadsheet inside it before translation"""

        logging.info('Loading the workbook.')
        try:
            self.workbook = load_workbook(filename=self.filename)
        except NameError:
            logging.error("Error occurred. Couldn't load the workbook. "
                          "Possibly wrong filename.")
            return None
        else:
            logging.info('Successfully loaded the workbook.')
            return [self.transl_sheet(sheet) for sheet in self.workbook]    # The iteration is here

    def save(self):

        """Saves the workbook into a new .xlsx file"""

        name = f'Translated - {self.filename}'
        logging.info('Saving the workbook into "%s"', name)
        self.workbook.save(filename=name)
        logging.info('Done.')


if __name__ == "__main__":
    spreadsheet = Spreadsheet(args.name, current=args.current, target=args.to)
    spreadsheet.translate()
    spreadsheet.save()
